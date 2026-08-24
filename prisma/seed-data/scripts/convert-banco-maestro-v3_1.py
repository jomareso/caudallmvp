# Registro histórico del script usado para convertir el Excel real
# "Banco_Maestro_Caudall_v3_1.xlsx" (corrección de la fundadora que agregó
# los 55 constructos y 110 variables de sesgos conductuales que faltaban) a
# prisma/seed-data/banco-maestro-v3.json. Requiere openpyxl (`pip install
# openpyxl`). No corre en CI ni en el flujo normal de seed — es un paso
# manual de una sola vez por revisión del Excel, igual que las conversiones
# anteriores (ver README.md). Para una futura corrección del Excel, ajustar
# SRC abajo y re-ejecutar; el diff contra el JSON anterior (ver historial de
# git) es la forma de confirmar que no se perdió ni cambió nada sin querer.

import sys
import json
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else 'Banco_Maestro_Caudall_v3_1.xlsx'
OUT = sys.argv[2] if len(sys.argv) > 2 else '../banco-maestro-v3.json'

wb = openpyxl.load_workbook(SRC, data_only=True)

warnings = []


def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ('—', '-', ''):
        return None
    return s


def rows_of(name):
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    data = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    return idx, data


def dim_or_none(raw):
    v = norm(raw)
    return None if v == 'BEHAVIORAL' else v


def yes_no(raw):
    v = norm(raw)
    if v is None:
        return None
    return v.strip().upper() in ('YES', 'SÍ', 'SI', 'TRUE')


BASE_PRIORITY_MAP = {'HIGH': 90, 'MEDIUM': 50, 'LOW': 10}
BURDEN_MAP = {'LOW': 1, 'MEDIUM': 3, 'HIGH': 5}

# ---- CONSTRUCTOS ----
cidx, crows = rows_of('CONSTRUCTOS')
constructs = []
for r in crows:
    weight_status = norm(r[cidx['Estado peso']])
    constructs.append({
        'code': norm(r[cidx['Código']]),
        'dimension': dim_or_none(r[cidx['Dimensión']]),
        'name': norm(r[cidx['Nombre']]),
        'definition': norm(r[cidx['Definición']]),
        'weightWithinDimension': float(r[cidx['Peso dentro de CFHI %']] or 0),
        'weightStatus': weight_status,
        'status': norm(r[cidx['Estado']]),
        'contributesToCfhi': weight_status != 'NO_CFHI'
    })

# ---- VARIABLES ----
vidx, vrows = rows_of('VARIABLES')
variables = []
for r in vrows:
    states_raw = norm(r[vidx['Estados posibles']])
    variables.append({
        'code': norm(r[vidx['Variable']]),
        'dimension': dim_or_none(r[vidx['Dimensión']]),
        'construct': norm(r[vidx['Constructo']]),
        'rawType': norm(r[vidx['Tipo']]),
        'states': states_raw.split('|') if states_raw else [],
        'affectsCfhiNote': norm(r[vidx['Afecta CFHI']]),
        'description': norm(r[vidx['Descripción']])
    })

# ---- OPCIONES (agrupadas por pregunta) ----
oidx, orows = rows_of('OPCIONES')
options_by_question = {}
for r in orows:
    qid = norm(r[oidx['ID Pregunta']])
    score_raw = r[oidx['Scoring 0-100']]
    options_by_question.setdefault(qid, []).append({
        'order': int(r[oidx['Orden']]),
        'text': norm(r[oidx['Texto opción']]),
        'state': norm(r[oidx['Estado resultante']]),
        'score': float(score_raw) if isinstance(score_raw, (int, float)) else None,
        'secondaryUpdates': norm(r[oidx['Actualizaciones secundarias']]),
        'friction': norm(r[oidx['Friction']]),
        'nextCandidates': norm(r[oidx['Next candidates']]),
        'notes': norm(r[oidx['Notas']])
    })

# ---- PREGUNTAS ----
pidx, prows = rows_of('PREGUNTAS')
questions = []
for r in prows:
    qid = norm(r[pidx['ID']])

    base_priority_raw = r[pidx['Base Priority']]
    if isinstance(base_priority_raw, str) and base_priority_raw.strip().upper() in BASE_PRIORITY_MAP:
        base_priority = BASE_PRIORITY_MAP[base_priority_raw.strip().upper()]
    elif isinstance(base_priority_raw, (int, float)):
        # Anomalía de datos: BEH-04 trae 0.9 en vez de HIGH/MEDIUM/LOW. El
        # banco v3.0 (ya en producción) tiene esta misma pregunta con
        # basePriority=90 — 0.9*100 coincide exacto, así que es la misma
        # intención (HIGH) mal tecleada en esta columna, no un valor nuevo.
        base_priority = round(base_priority_raw * 100)
        warnings.append(f'{qid}: Base Priority trae {base_priority_raw!r} (numérico, no HIGH/MEDIUM/LOW) — se interpretó como {base_priority} (coincide con v3.0)')
    else:
        warnings.append(f'{qid}: Base Priority vacío o irreconocible {base_priority_raw!r}, se usó MEDIUM (50)')
        base_priority = 50

    burden_raw = norm(r[pidx['Burden']])
    burden = BURDEN_MAP.get((burden_raw or '').upper(), 1)
    if burden_raw and burden_raw.upper() not in BURDEN_MAP:
        warnings.append(f'{qid}: Burden inesperado {burden_raw!r}, se usó LOW (1)')

    ask_if = norm(r[pidx['ASK_IF']])
    if ask_if and ask_if.strip().upper() == 'TRUE':
        ask_if = None

    skip_if = norm(r[pidx['SKIP_IF']])
    # BEH-04: SKIP_IF trae 'MEDIUM' (valor de otra columna, desplazado) — no
    # es una condición válida, se descarta en vez de cargarla como regla.
    if skip_if and skip_if.strip().upper() in ('LOW', 'MEDIUM', 'HIGH'):
        warnings.append(f'{qid}: SKIP_IF tenía {skip_if!r} (no es una condición) — se descartó')
        skip_if = None

    min_conf_raw = r[pidx['Min confidence skip']]
    min_conf = round(float(min_conf_raw) * 100) if isinstance(min_conf_raw, (int, float)) else None

    q = {
        'id': qid,
        'dimension': dim_or_none(r[pidx['Dimensión']]),
        'construct': norm(r[pidx['Constructo']]),
        'variable': norm(r[pidx['Variable objetivo']]),
        'primaryOwnerConstruct': norm(r[pidx['Primary Owner']]),
        'textUX': norm(r[pidx['Pregunta UX']]),
        'whyAsk': norm(r[pidx['WHY_ASK']]),
        'role': norm(r[pidx['Tipo']]),
        'anchor': yes_no(r[pidx['Ancla']]) or False,
        'askIfRaw': ask_if,
        'skipIfRaw': skip_if,
        'basePriority': base_priority,
        'informationValue': float(r[pidx['Information Value']] or 0),
        'scoringValue': float(r[pidx['Scoring Value']] or 0),
        'routingValue': float(r[pidx['Routing Value']] or 0),
        'safetyValue': float(r[pidx['Safety Value']] or 0),
        'rootCauseValue': float(r[pidx['Root Cause Value']] or 0),
        'uncertaintyReduction': float(r[pidx['Uncertainty Reduction']] or 0),
        'burden': burden,
        'inferenceSubstitutionAllowed': yes_no(r[pidx['Inference substitution']]) or False,
        'minConfidenceToSkip': min_conf,
        'frictionTarget': norm(r[pidx['Friction target']]),
        'aiRegenerationAllowed': yes_no(r[pidx['AI regeneration']]) or False,
        'coreLogicEditable': yes_no(r[pidx['CORE logic editable']]) or False,
        'status': norm(r[pidx['Estado']]),
        'version': norm(r[pidx['Versión']]),
        'benchmarkSource': norm(r[pidx['Benchmark principal']]),
        'methodologicalFunction': norm(r[pidx['Función metodológica']]),
        'behavioralConstruct': norm(r[pidx['Constructo conductual']]),
        'options': sorted(options_by_question.get(qid, []), key=lambda o: o['order'])
    }
    questions.append(q)

# ---- Validación de referencias (debe quedar en 0 pendientes) ----
construct_codes = {c['code'] for c in constructs}
variable_codes = {v['code'] for v in variables}
questions_pending_ids = []
for q in questions:
    if q['variable'] not in variable_codes:
        questions_pending_ids.append(q['id'])
        warnings.append(f"{q['id']}: variable {q['variable']} no existe en VARIABLES")

# ---- INFERENCIAS (STRONG/WEAK -> inferenceRules, FORBIDDEN -> forbiddenInferences) ----
iidx, irows = rows_of('INFERENCIAS')
inference_rules = []
forbidden_inferences = []
for r in irows:
    tipo = norm(r[iidx['Tipo']])
    affected_raw = norm(r[iidx['Preguntas afectadas']])
    affected = [x.strip() for x in affected_raw.split(',')] if affected_raw else []
    if tipo == 'FORBIDDEN':
        cond = norm(r[iidx['Condición fuente']]) or ''
        if '=' not in cond:
            warnings.append(f"{r[iidx['ID']]}: condición FORBIDDEN sin '=': {cond!r}")
            continue
        source_var, source_val = [p.strip() for p in cond.split('=', 1)]
        forbidden_inferences.append({
            'sourceVariableCode': source_var,
            'sourceValue': source_val,
            'targetVariableCode': norm(r[iidx['Variable destino']]),
            'targetValue': norm(r[iidx['Valor destino']]),
            'reason': norm(r[iidx['Notas']])
        })
    else:
        inference_rules.append({
            'code': norm(r[iidx['ID']]),
            'type': tipo,
            'sourceConditionRaw': norm(r[iidx['Condición fuente']]),
            'targetVariableCode': norm(r[iidx['Variable destino']]),
            'targetValue': norm(r[iidx['Valor destino']]),
            'confidence': float(r[iidx['Confidence']] or 0),
            'canSubstituteQuestion': yes_no(r[iidx['Puede sustituir pregunta']]) or False,
            'affectedQuestionCodes': affected,
            'notes': norm(r[iidx['Notas']])
        })

# ---- QA ----
qidx, qrows = rows_of('QA')
qa_scenarios = []
for r in qrows:
    qa_scenarios.append({
        'code': norm(r[qidx['ID']]),
        'scenario': norm(r[qidx['Escenario']]),
        'precondition': norm(r[qidx['Input / precondición']]),
        'expectedResult': norm(r[qidx['Resultado esperado']]),
        'severity': norm(r[qidx['Severidad']])
    })

# ---- BEHAVIORAL & COPY ----
# BehavioralTechnique.avoidWhen es NOT NULL en el schema (a diferencia de
# la mayoría de estos campos) — COPY_CORE/COPY_MOBILE (guía de estilo, no
# técnicas conductuales reales) traen "—" en "No usar cuando" porque no
# aplica, así que se conserva el texto tal cual en vez de normalizar a
# null (igual que ya estaba guardado en producción antes de este cambio).
bidx, brows = rows_of('BEHAVIORAL & COPY')
behavioral_techniques = []
for r in brows:
    behavioral_techniques.append({
        'frictionCode': norm(r[bidx['Fricción / estado']]),
        'technique': norm(r[bidx['Técnica candidata']]),
        'useWhen': norm(r[bidx['Usar cuando']]),
        'avoidWhen': norm(r[bidx['No usar cuando']]) or '—',
        'copyTransformation': norm(r[bidx['Transformación UX / copy']]),
        'example': norm(r[bidx['Ejemplo']])
    })

# ---- MAPA CONDUCTUAL ----
midx, mrows = rows_of('MAPA CONDUCTUAL')
behavioral_bias_map = []
for r in mrows:
    behavioral_bias_map.append({
        'construct': norm(r[midx['Constructo']]),
        'whatItDetects': norm(r[midx['Qué busca detectar']]),
        'whenToAsk': norm(r[midx['Cuándo preguntar']]),
        'whatNotToConclude': norm(r[midx['Qué NO concluir']]),
        'candidateIntervention': norm(r[midx['Intervención candidata']]),
        'benchmarks': norm(r[midx['Benchmarks']])
    })

out = {
    # Se mantiene 3.0.0 (no se sube a 3.1.0) a propósito: varios motores
    # hacen `methodology.findFirst({ where: { status: 'ACTIVE' } })` sin
    # ordenar por versión — si esto creara una Methodology nueva, quedarían
    # dos filas ACTIVE a la vez y ese findFirst sería no determinista. Esta
    # sigue siendo la misma metodología, solo con las preguntas que
    # faltaban ya cargadas, así que actualiza la fila existente en vez de
    # crear una nueva.
    'methodologyVersion': '3.0.0',
    'questionBankVersion': '3.0.0',
    'constructs': constructs,
    'variables': variables,
    'questions': questions,
    'questionsPendingIds': questions_pending_ids,
    'inferenceRules': inference_rules,
    'forbiddenInferences': forbidden_inferences,
    'qaScenarios': qa_scenarios,
    'behavioralTechniques': behavioral_techniques,
    'behavioralBiasMap': behavioral_bias_map
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f'constructs={len(constructs)} variables={len(variables)} questions={len(questions)} pending={len(questions_pending_ids)}')
print(f'inferenceRules={len(inference_rules)} forbiddenInferences={len(forbidden_inferences)} qaScenarios={len(qa_scenarios)}')
print(f'behavioralTechniques={len(behavioral_techniques)} behavioralBiasMap={len(behavioral_bias_map)}')
print()
print(f'{len(warnings)} advertencias:')
for w in warnings:
    print(' -', w)
