# Registro histórico del script usado para convertir el Excel real
# "Banco_Maestro_Caudall_v4_1_SIMPLIFICADO.xlsx". Misma estructura de 4
# pestañas que v3.6/v4.0 (00_ARQUITECTURA no se carga) — sin cambios de
# constructos ni de columnas. El único cambio real de v4.1 frente a v3.6/
# v4.0 es la columna `Estado` de 02_BANCO_PREGUNTAS: bajó de 211 a 101
# preguntas `ACTIVA`, y las inactivas ahora se llaman `RESERVA` en vez de
# `DRAFT` — `toQuestionStatus()` en sync-banco-maestro.ts ya trata
# cualquier valor que no sea literalmente 'ACTIVA' como DRAFT, así que
# `RESERVA` cae ahí sin tocar código. Es la simplificación pedida por la
# fundadora (menos preguntas activas), lograda solo con el Excel — la capa
# de constructos (55 sesgos × 5 dimensiones) sigue intacta, no es el
# rediseño de la "auditoría metodológica" (secciones F-J del Excel, que
# siguen siendo una propuesta aparte, no aplicada al banco todavía).
#
# 01_METODOLOGIA y 03_REGLAS_QA también crecieron con secciones nuevas
# (E-J: matriz de validación, auditoría, plan de validación, contexto/
# segmentación) que este script deliberadamente NO lee — están fuera de
# los rangos de fila hardcodeados de abajo, igual que en v3.6/v4.0. La
# única adaptación real fue en la sección de escenarios QA: v4.1 le
# agregó una subsección "C. CONTEXTO..." con su propio título y fila de
# encabezado repetida en medio de la lista de QA-*, que si no se filtra
# se cuela como un escenario QA falso (ver filtro más abajo).
#
# Requiere openpyxl. No corre en CI — paso manual de una sola vez por
# revisión del Excel (ver README.md).

import sys
import json
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else 'Banco_Maestro_Caudall_v4_1_SIMPLIFICADO.xlsx'
OUT = sys.argv[2] if len(sys.argv) > 2 else '../banco-maestro-v3.json'

wb = openpyxl.load_workbook(SRC, data_only=True)
warnings = []


def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ('—', '-', ''):
        return None
    return s


def dim_or_none(raw):
    v = norm(raw)
    return None if v == 'BEHAVIORAL' else v


def yes_no(raw):
    v = norm(raw)
    if v is None:
        return None
    return v.strip().upper() in ('YES', 'SÍ', 'SI', 'TRUE')


def section_rows(sheet, header_row, last_row):
    ws = wb[sheet]
    headers = [c.value for c in ws[header_row]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = [r for r in ws.iter_rows(min_row=header_row + 1, max_row=last_row, values_only=True) if r[0] is not None]
    return idx, rows


BASE_PRIORITY_MAP = {'HIGH': 90, 'MEDIUM': 50, 'LOW': 10}
BURDEN_MAP = {'LOW': 1, 'MEDIUM': 3, 'HIGH': 5}

# ---- 01_METODOLOGIA: A. CONSTRUCTOS (filas 2-79) ----
cidx, crows = section_rows('01_METODOLOGIA', 2, 79)
constructs = []
for r in crows:
    weight_status = norm(r[cidx['Estado peso']])
    constructs.append({
        'code': norm(r[cidx['Código']]),
        'dimension': dim_or_none(r[cidx['Dimensión']]),
        'name': norm(r[cidx['Nombre']]),
        'definition': norm(r[cidx['Definición']]),
        'weightWithinDimension': float(r[cidx['Peso dentro de CFHI %']] or 0),
        'weightStatus': weight_status,
        'status': norm(r[cidx['Estado']]),
        'contributesToCfhi': weight_status != 'NO_CFHI'
    })

# ---- 01_METODOLOGIA: B. VARIABLES (filas 81-257) ----
vidx, vrows = section_rows('01_METODOLOGIA', 81, 257)
variables = []
for r in vrows:
    states_raw = norm(r[vidx['Estados posibles']])
    variables.append({
        'code': norm(r[vidx['Variable']]),
        'dimension': dim_or_none(r[vidx['Dimensión']]),
        'construct': norm(r[vidx['Constructo']]),
        'rawType': norm(r[vidx['Tipo']]),
        'states': states_raw.split('|') if states_raw else [],
        'affectsCfhiNote': norm(r[vidx['Afecta CFHI']]),
        'description': norm(r[vidx['Descripción']])
    })

# ---- 01_METODOLOGIA: C. MAPA CONDUCTUAL (filas 259-272) ----
midx, mrows = section_rows('01_METODOLOGIA', 259, 272)
behavioral_bias_map = []
for r in mrows:
    behavioral_bias_map.append({
        'construct': norm(r[midx['Constructo']]),
        'whatItDetects': norm(r[midx['Qué busca detectar']]),
        'whenToAsk': norm(r[midx['Cuándo preguntar']]),
        'whatNotToConclude': norm(r[midx['Qué NO concluir']]),
        'candidateIntervention': norm(r[midx['Intervención candidata']]),
        'benchmarks': norm(r[midx['Benchmarks']])
    })

# ---- 01_METODOLOGIA: D. BEHAVIORAL & COPY (filas 274-287) ----
bidx, brows = section_rows('01_METODOLOGIA', 274, 287)
behavioral_techniques = []
for r in brows:
    behavioral_techniques.append({
        'frictionCode': norm(r[bidx['Fricción / estado']]),
        'technique': norm(r[bidx['Técnica candidata']]),
        'useWhen': norm(r[bidx['Usar cuando']]),
        'avoidWhen': norm(r[bidx['No usar cuando']]) or '—',
        'copyTransformation': norm(r[bidx['Transformación UX / copy']]),
        'example': norm(r[bidx['Ejemplo']])
    })

# ---- 03_REGLAS_QA: A. INFERENCIAS Y REGLAS PROHIBIDAS (filas 2-21) ----
iidx, irows = section_rows('03_REGLAS_QA', 2, 21)
inference_rules = []
forbidden_inferences = []
for r in irows:
    tipo = norm(r[iidx['Tipo']])
    affected_raw = norm(r[iidx['Preguntas afectadas']])
    affected = [x.strip() for x in affected_raw.split(',')] if affected_raw else []
    if tipo == 'FORBIDDEN':
        cond = norm(r[iidx['Condición fuente']]) or ''
        if '=' not in cond:
            warnings.append(f"{r[iidx['ID']]}: condición FORBIDDEN sin '=': {cond!r}")
            continue
        source_var, source_val = [p.strip() for p in cond.split('=', 1)]
        forbidden_inferences.append({
            'sourceVariableCode': source_var,
            'sourceValue': source_val,
            'targetVariableCode': norm(r[iidx['Variable destino']]),
            'targetValue': norm(r[iidx['Valor destino']]),
            'reason': norm(r[iidx['Notas']])
        })
    else:
        inference_rules.append({
            'code': norm(r[iidx['ID']]),
            'type': tipo,
            'sourceConditionRaw': norm(r[iidx['Condición fuente']]),
            'targetVariableCode': norm(r[iidx['Variable destino']]),
            'targetValue': norm(r[iidx['Valor destino']]),
            'confidence': float(r[iidx['Confidence']] or 0),
            'canSubstituteQuestion': yes_no(r[iidx['Puede sustituir pregunta']]) or False,
            'affectedQuestionCodes': affected,
            'notes': norm(r[iidx['Notas']])
        })

# ---- 03_REGLAS_QA: B. QA / PRUEBAS DE PUBLICACIÓN (desde fila 25) ----
# v4.1 agrega una subsección "C. CONTEXTO, SEGMENTACIÓN..." (fila 58) con
# su propio título y fila de encabezado repetida (fila 59) en medio de la
# lista de QA-*, antes de que retomen QA-033..044 — se descartan ambas
# filas para no crear un escenario QA falso.
ws_qa = wb['03_REGLAS_QA']
qidx = {h: i for i, h in enumerate([c.value for c in ws_qa[25]])}
SECTION_TITLE_PREFIXES = tuple(f'{c}. ' for c in 'ABCDEFGHIJ')
qrows = [
    r for r in ws_qa.iter_rows(min_row=26, max_row=ws_qa.max_row, values_only=True)
    if r[0] and str(r[0]) != 'ID' and not str(r[0]).startswith(SECTION_TITLE_PREFIXES)
]
qa_scenarios = []
for r in qrows:
    qa_scenarios.append({
        'code': norm(r[qidx['ID']]),
        'scenario': norm(r[qidx['Escenario']]),
        'precondition': norm(r[qidx['Input / precondición']]),
        'expectedResult': norm(r[qidx['Resultado esperado']]),
        'severity': norm(r[qidx['Severidad']])
    })

# ---- 02_BANCO_PREGUNTAS (una fila por opción; agrupar por QUESTION_ID) ----
pws = wb['02_BANCO_PREGUNTAS']
pheaders = [c.value for c in pws[1]]
pidx = {h: i for i, h in enumerate(pheaders)}
all_rows = [r for r in pws.iter_rows(min_row=2, values_only=True) if r[0]]

questions_by_id = {}
options_by_question = {}
for r in all_rows:
    qid = norm(r[pidx['QUESTION_ID']])

    if qid not in questions_by_id:
        base_priority_raw = r[pidx['Base Priority']]
        if isinstance(base_priority_raw, str) and base_priority_raw.strip().upper() in BASE_PRIORITY_MAP:
            base_priority = BASE_PRIORITY_MAP[base_priority_raw.strip().upper()]
        elif isinstance(base_priority_raw, (int, float)):
            base_priority = round(base_priority_raw * 100)
            warnings.append(f'{qid}: Base Priority trae {base_priority_raw!r} (numérico) — se interpretó como {base_priority}')
        else:
            warnings.append(f'{qid}: Base Priority vacío o irreconocible {base_priority_raw!r}, se usó MEDIUM (50)')
            base_priority = 50

        burden_raw = norm(r[pidx['Burden']])
        burden = BURDEN_MAP.get((burden_raw or '').upper(), 1)
        if burden_raw and burden_raw.upper() not in BURDEN_MAP:
            warnings.append(f'{qid}: Burden inesperado {burden_raw!r}, se usó LOW (1)')

        ask_if = norm(r[pidx['ASK_IF']])
        if ask_if and ask_if.strip().upper() == 'TRUE':
            ask_if = None

        skip_if = norm(r[pidx['SKIP_IF']])
        if skip_if and skip_if.strip().upper() in ('LOW', 'MEDIUM', 'HIGH'):
            warnings.append(f'{qid}: SKIP_IF tenía {skip_if!r} (no es una condición) — se descartó')
            skip_if = None

        min_conf_raw = r[pidx['Min confidence skip']]
        min_conf = round(float(min_conf_raw) * 100) if isinstance(min_conf_raw, (int, float)) else None

        questions_by_id[qid] = {
            'id': qid,
            'dimension': dim_or_none(r[pidx['Dimensión']]),
            'construct': norm(r[pidx['Constructo']]),
            'variable': norm(r[pidx['Variable objetivo']]),
            'primaryOwnerConstruct': norm(r[pidx['Primary Owner']]),
            'textUX': norm(r[pidx['Pregunta UX']]),
            'whyAsk': norm(r[pidx['WHY_ASK']]),
            'role': norm(r[pidx['Tipo']]),
            'anchor': yes_no(r[pidx['Ancla']]) or False,
            'askIfRaw': ask_if,
            'skipIfRaw': skip_if,
            'basePriority': base_priority,
            'informationValue': float(r[pidx['Information Value']] or 0),
            'scoringValue': float(r[pidx['Scoring Value']] or 0),
            'routingValue': float(r[pidx['Routing Value']] or 0),
            'safetyValue': float(r[pidx['Safety Value']] or 0),
            'rootCauseValue': float(r[pidx['Root Cause Value']] or 0),
            'uncertaintyReduction': float(r[pidx['Uncertainty Reduction']] or 0),
            'burden': burden,
            'inferenceSubstitutionAllowed': yes_no(r[pidx['Inference substitution']]) or False,
            'minConfidenceToSkip': min_conf,
            'frictionTarget': norm(r[pidx['Friction target']]),
            'aiRegenerationAllowed': yes_no(r[pidx['AI regeneration']]) or False,
            'coreLogicEditable': yes_no(r[pidx['CORE logic editable']]) or False,
            'status': norm(r[pidx['Estado']]),
            'version': norm(r[pidx['Versión']]),
            'benchmarkSource': norm(r[pidx['Benchmark principal']]),
            'methodologicalFunction': norm(r[pidx['Función metodológica']]),
            'behavioralConstruct': norm(r[pidx['Constructo conductual']])
        }
        options_by_question[qid] = []

    score_raw = r[pidx['Opción scoring 0-100']]
    order_raw = r[pidx['Opción orden']]
    options_by_question[qid].append({
        'order': int(order_raw) if order_raw is not None else len(options_by_question[qid]) + 1,
        'text': norm(r[pidx['Opción texto']]),
        'state': norm(r[pidx['Opción estado/resultante']]),
        'score': float(score_raw) if isinstance(score_raw, (int, float)) else None,
        'secondaryUpdates': norm(r[pidx['Opción actualizaciones secundarias']]),
        'friction': norm(r[pidx['Opción friction']]),
        'nextCandidates': norm(r[pidx['Opción next candidates']]),
        'notes': norm(r[pidx['Opción notas']])
    })

questions = []
for qid, q in questions_by_id.items():
    q['options'] = sorted(options_by_question[qid], key=lambda o: o['order'])
    questions.append(q)

# ---- Completar variables de contexto que faltan en la pestaña B. VARIABLES ----
# Las 7 preguntas CTX-01..07 (bloque de contexto del documento de
# metodología v1.5: edad, dependientes, responsabilidad del hogar,
# ingreso, patrón de ingreso, educación, y el opt-in de comparación) están
# completas con sus opciones en 02_BANCO_PREGUNTAS, pero la fila de su
# variable objetivo no se agregó a la sección B — mismo tipo de vacío que
# documentó el propio banco al pasar de v3.0 a v3.1 (220 preguntas sin su
# fila de VARIABLES). Se deriva aquí mecánicamente de las opciones ya
# presentes, sin inventar contenido nuevo — todas apuntan al constructo
# CTX_SEGMENTATION (No CFHI), que ya existe en la sección A.
construct_codes_present = {c['code'] for c in constructs}
variable_codes_present = {v['code'] for v in variables}
for q in questions:
    if q['dimension'] != 'CONTEXTO':
        continue
    if q['variable'] in variable_codes_present:
        continue
    if 'CTX_SEGMENTATION' not in construct_codes_present:
        warnings.append(f"{q['id']}: no se pudo completar {q['variable']} — falta el constructo CTX_SEGMENTATION")
        continue
    states = [o['state'] for o in sorted(options_by_question[q['id']], key=lambda o: o['order']) if o['state']]
    variables.append({
        'code': q['variable'],
        'dimension': 'CONTEXTO',
        'construct': 'CTX_SEGMENTATION',
        'rawType': 'CONTEXT',
        'states': states,
        'affectsCfhiNote': 'NO_CFHI',
        'description': f"Completada automáticamente desde las opciones de {q['id']} (ver comentario en el script)."
    })
    variable_codes_present.add(q['variable'])
    warnings.append(f"{q['variable']}: variable completada automáticamente desde las opciones de {q['id']} (no estaba en VARIABLES)")

# ---- Validación de referencias ----
construct_codes = {c['code'] for c in constructs}
variable_codes = {v['code'] for v in variables}
questions_pending_ids = []
for q in questions:
    if q['variable'] not in variable_codes:
        questions_pending_ids.append(q['id'])
        warnings.append(f"{q['id']}: variable {q['variable']} no existe en VARIABLES")

out = {
    'methodologyVersion': '3.0.0',
    'questionBankVersion': '3.0.0',
    'constructs': constructs,
    'variables': variables,
    'questions': questions,
    'questionsPendingIds': questions_pending_ids,
    'inferenceRules': inference_rules,
    'forbiddenInferences': forbidden_inferences,
    'qaScenarios': qa_scenarios,
    'behavioralTechniques': behavioral_techniques,
    'behavioralBiasMap': behavioral_bias_map
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f'constructs={len(constructs)} variables={len(variables)} questions={len(questions)} pending={len(questions_pending_ids)}')
print(f'inferenceRules={len(inference_rules)} forbiddenInferences={len(forbidden_inferences)} qaScenarios={len(qa_scenarios)}')
print(f'behavioralTechniques={len(behavioral_techniques)} behavioralBiasMap={len(behavioral_bias_map)}')
print()
print(f'{len(warnings)} advertencias:')
for w in warnings:
    print(' -', w)
